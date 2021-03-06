# -*- coding: utf-8 -*-
"""
Created on Wed Jun  5 20:21:57 2019

@author: 文明研究员
"""
import numpy as np
import matplotlib.pyplot as plt
from const import Const
from scipy.integrate import odeint
from openpyxl import Workbook,load_workbook
from valueDataBase import DataBase

class Test(object):
    con=Const(0.3)
    values=[]
    dValues=[]
    dr=0.0
    num=1000
    f=[]
    @classmethod
    def ode_fun_B(cls,Z,r,c_P,c_T,c_M,others,ST):
        P,T,M,G,DG,L=Z
        #if L<-1000:
            #raise LOutOfRangeError('L<-1000')
        g,alpha,beta,Lambda,R_B=others

        dP=c_P*M*P/(T*r**2)
        P_0=cls.con.P_0
        dT=min(c_T*1e24*abs(L)*P**(alpha+1)*T**(beta-4)/M,g)*(T*dP/P)
        dM=c_M*r**2*P/T
        dG=DG
        if P<=0.0:
            print("Error")
        dL=Lambda*7.15e-5*(dG**2+G**2/r**2)/(9e9*10**cls.f(np.log10(P*P_0)-6)*R_B)
        if dT==g*(T*dP/P):
            dL+=1e-24*cls.con.M_e*cls.con.T_0*dM*ST
        else:
            dL+=0
        ddG=6*G/r**2+(0.5*1e3*(cls.f(np.log10(P*P_0)-6+1e-3)-cls.f(np.log10(P*P_0)-6-1e-3))*dP/P)*dG+(r>cls.con.depth)*2*cls.con.M_v*(1/r-cls.con.depth**2/r**3)/cls.con.c**2/cls.con.R_B

        return [dP,dT,dM,dG,ddG,dL]

    @classmethod
    def find_RCB_index(cls):
        P,T,M,L=cls.P,cls.T,cls.M,cls.L
        judge=(cls.con.c_T*1e24*abs(L)*P**(cls.con.alpha+1)*T**(cls.con.beta-4)/M)<(np.ones(len(L))*cls.con.g_ad)
        for i in range(len(judge)):
            if judge[i]!=judge[0]:
                cls.RCB_index=i
                return i 
        return 0

    @classmethod
    def cal_ML_simple_B(cls,ST,L_s,B=False,G=1e-12,dG=1e-15):
        c_P,c_T,c_M=cls.con.cal_const()
        g,alpha,beta=cls.con.g_ad,cls.con.alpha,cls.con.beta
        R_in,R_out=cls.con.R_p,cls.con.R_out
        num=cls.num
        base=DataBase()
        base.creat_sigma()
        cls.f=base.f_sigma

        r=np.linspace(1,R_in/R_out,num)
        cls.dr=r[0]-r[1]
        if B:
            initial=(1.,1.,cls.con.M_p,G,dG,L_s)
        else:
            initial=(1.,1.,cls.con.M_p,0.0,0.0,L_s)
        others=(g,alpha,beta,cls.con.Lambda,cls.con.R_B)
        result=odeint(cls.ode_fun_B,initial,r,args=(c_P,c_T,c_M,others,ST))
        P,T,M,G,dg,L=result[:,0],result[:,1],result[:,2],result[:,3],result[:,4],result[:,5]
        cls.P,cls.T,cls.M,cls.L=P,T,M,L
        cls.r,cls.g=r,G
        cls.dg=dg

        cls.find_RCB_index()
        return M[-1],L[-1],G[-1]

    @classmethod
    def sigma(cls,P,T):
        return cls.con.sigma_1*T**(3./4)*P**(-1./2)*np.exp(-cls.con.sigma_2/T)

    @classmethod
    def find_ML_simple(cls,targets, inits, steps, errors):
        """The order of targets is [M_c, L_c], of inits is [ST, L_s](Note that the unit of L_s is 1e+24 erg/s), of steps and error is the same as inits and targets."""
        """The way to find the result is one-dimention Newton iteration. First adjustion is L_s and then is ST."""
        ST,L_s=inits
        M_c_need,L_c_need=targets
        error_M_c, error_L_c=errors
        M_c,L_c=cls.cal_ML_simple_B(ST,L_s,True,0.0,-1)[0:2]
        while abs(M_c-M_c_need)>error_M_c or abs(L_c-L_c_need)>error_L_c:
            while abs(M_c-M_c_need)>error_M_c:
                point1=(L_s,M_c)
                point2=(L_s+L_s*steps[1],cls.cal_ML_simple_B(ST,L_s+L_s*steps[1],True,0.0,-1)[0])
                k=cls.grad(point1,point2)
                if k==0:
                    raise ZeroDivisionError('k is zero.Error is M_c,M_p is {0}'.format(cls.con.M_p))
                else:
                    L_s=(M_c_need-M_c)/k+L_s
                    if L_s<0:
                        L_s=-L_s/10
                    M_c,L_c=cls.cal_ML_simple_B(ST,L_s,True,0.0,-1)[0:2]
            print(M_c,L_c)
            while abs(L_c_need-L_c)>error_L_c:
                ST_next=ST+steps
                L_c_next=cls.cal_ML_simple_B(ST_next,L_s,True,0.0,-1)[1]
                point1=[ST,L_c]
                point2=[ST_next,L_c_next]
                k=cls.grad(point1,point2)
                if k==0:
                    raise ZeroDivisionError('k is zero.Error is L_c. M_p is {0}'.format(cls.con.M_p))
                else:
                    ST=(L_c_need-L_c)/k+ST
                    L_c=cls.cal_ML_simple_B(ST,L_s,True,0.0,-1)[1]
            print(M_c,L_c)
        return ST,L_s

    @classmethod
    def grad(cls,point1,point2):
        """Return the grad of two points, so for one-dimention Newton iteration."""
        return (point2[1]-point1[1])/(point2[0]-point1[0])
    
    @classmethod
    def creat_test_data(cls,fileName='TestData.xlsx'):
        wb=Workbook()
        wb.create_sheet('result',index=0)
        sheet=wb[wb.sheetnames[0]]
        length=len(cls.values)
        sheet.append(['Information: ','r_range=linspace({0},{1},{2})'.format(cls.values[0][0],cls.values[length-1][0], length),'M_c={0}M_e'.format(cls.con.M_e),'M_p={0}M_e'.format(cls.con.M_p),'a={0}AU'.format(cls.con.a)])
        sheet.append(['r','P','T','M','G','DG','L'])
        for i in range(len(cls.values)):
            sheet.append(cls.values[i])
        sheet.append(['dr','dP','dT','dM','dG','dDG','dL'])
        for i in range(len(cls.values)):
            sheet.append(cls.dValues[i])
        wb.save(fileName)

    @classmethod
    def creat_t(cls,initValues,fileName='TestT.xlsx'):
        """The order of initValues is M_p, ST, L_s"""
        M_p,ST,L_s=initValues
        P=np.ones(len(M_p))
        T=np.ones(len(M_p))
        L=np.ones(len(M_p))
        for order in range(len(M_p)):
            cls.con.set_M_p(M_p[order])
            ST[order]=cls.find_ML_simple(0.0,[ST[order],L_s[order]],1e-7,1e-3)[0]
            print("Finished: {0}/200".format(order))
            P[order]=cls.P[-1]
            T[order]=cls.T[-1]
            L[order]=cls.L[0]
        cls.ST=ST
        cls.t=np.ones(len(M_p))
        n_0=cls.con.R/cls.con.mu
        C_p=5./2.
        cls.t[0]=0.0
        for i in range(len(M_p)-1):
            T_1,T_2=cls.T[i],cls.T[i+1]
            P_1,P_2=cls.P[i],cls.P[i+1]
            ST_1,ST_2=cls.ST[i],cls.ST[i+1]
            delta_1=C_p*(1/T_2+1/T_1)*(T_2-T_1)
            delta_2=(1/P_2+1/P_1)*(P_2-P_1)
            dt=n_0*(delta_1+delta_2)/((ST_1+ST_2)*cls.con.Myr)
            cls.t[i+1]=cls.t[i]+dt

        wb=Workbook()
        wb.create_sheet('result',index=0)
        sheet=wb[wb.sheetnames[0]]
        length=len(M_p)
        sheet.append(['Information: ','range=linspace({0},{1},{2})'.format(M_p[0],M_p[length-1], length),'M_c={0}M_e'.format(cls.con.M_e),'a={0}AU'.format(cls.con.a)])
        sheet.append(['M_p/M_e','ST','L/e+24erg/s','t/Myr'])
        for i in range(len(M_p)):
            sheet.append([M_p[i],cls.ST[i],cls.L[i],cls.t[i]])
        wb.save(fileName)
