'''
@Author: Xiao Liang
@Date: 2019-07-26 10:51:33
@LastEditors: Xiao Liang
@LastEditTime: 2019-09-22 21:12:03
@Description: file content
'''
from const import Const
import numpy as np
import warnings
from scipy.integrate import odeint
from openpyxl import Workbook,load_workbook

class ToAdjust(Exception):
    def __init__(self):
        Exception.__init__(self,'None')

#Can not stop.
class Methods(object):
    con=Const()    #If the Const you need is not initial, you need to change it.
    P=[]
    T=[]
    M=[]
    L=[]
    test1=[]
    test2=[]
    g=[]
    dg=[]
    L_store=0
    message=[]
    judge=True
    @classmethod
    def ode_fun(cls,Z,r,c_P,c_T,c_M,others,ST):
        P,T,M,G,DG,L=Z
        if L<=-0.5:
            cls.L_test=L
            raise ToAdjust()
        g,alpha,beta,Lambda,R_B=others

        dP=c_P*M*P/(T*r**2)
        dT=min(c_T*1e24*abs(L)*P**(alpha+1)*T**(beta-4)/M,g)*(T*dP/P)

        dM=c_M*r**2*P/T

        dG=DG
        
        ddG=6*G/r**2+((-0.5*dP/P)+(0.75/T+cls.con.sigma_2/T**2)*dT)*dG
        
        dL=Lambda*7.15e-5*(dG**2+G**2/r**2)/(cls.sigma(P,T)*R_B)
        
        if dT==g*(T*dP/P):
            dL+=1e-24*cls.con.M_e*cls.con.T_0*dM*ST
        else:
            dL+=0
        cls.test1.append([dP,dT,dM,dG,ddG,dL])
        cls.test2.append([P,T,M,G,dG,L])
        return [dP,dT,dM,dG,ddG,dL]

    @classmethod
    def cal_ML_simple(cls,ST,L_s,B=False):
        c_P,c_T,c_M=cls.con.cal_const()
        g,alpha,beta=cls.con.g_ad,cls.con.alpha,cls.con.beta
        R_in,R_out=cls.con.R_p,cls.con.R_out
        num=800

        r=np.linspace(1,R_in/R_out,num)
        if B:
            initial=(1.,1.,cls.con.M_p,1e-12,1e-15,L_s)
        else:
            initial=(1.,1.,cls.con.M_p,0.0,0.0,L_s)
        others=(g,alpha,beta,cls.con.Lambda,cls.con.R_B)
        cls.test2=[]
        result=odeint(cls.ode_fun,initial,r,args=(c_P,c_T,c_M,others,ST))
        P,T,M,g,dg,L=result[:,0],result[:,1],result[:,2],result[:,3],result[:,4],result[:,5]
        cls.P,cls.T,cls.M,cls.L=P,T,M,L
        cls.r,cls.g=r,g
        cls.dg=dg
        return M[-1],L[-1]

    @classmethod
    def cal_L_safe(cls,ST,L_s):
        try:
            warnings.simplefilter('error')
            L=cls.cal_ML_simple(ST,L_s)[1]
        except:
            ST=cls.find_safe_point(ST,L_s)
            L=cls.cal_ML_simple(ST,L_s)[1]
        finally:
            return ST,L
    
    @classmethod
    def grad(cls,point1,point2):
        """Return the grad of two points, so for one-dimention Newton iteration."""
        return (point2[1]-point1[1])/(point2[0]-point1[0])
    
    @classmethod
    def sigma(cls,P,T):
        return cls.con.sigma_1*T**(3./4)*P**(-1./2)*np.exp(-cls.con.sigma_2/T)
    
    
    @classmethod
    def find_L_M(cls, targets, inits, steps, errors):
        """The order of targets is [M_c, L_c], of inits is [ST, L_s](Note that the unit of L_s is 1e+24 erg/s), of steps and error is the same as inits and targets."""
        """The way to find the result is one-dimention Newton iteration. First adjustion is L_s and then is ST."""
        ST,L_s=inits
        M_c_need,L_c_need=targets
        error_M_c, error_L_c=errors
        M_c,L_c=cls.cal_ML_simple(ST,L_s)
        while abs(M_c-M_c_need)>error_M_c or abs(L_c-L_c_need)>error_L_c:
            while abs(M_c-M_c_need)>error_M_c:
                point1=(L_s,M_c)
                point2=(L_s+steps[1],cls.cal_ML_simple(ST,L_s+steps[1])[0])
                k=cls.grad(point1,point2)
                if k==0:
                    raise ZeroDivisionError('k is zero.')
                else:
                    L_s=(M_c_need-M_c)/k+L_s
                    if L_s<0:
                        L_s=-L_s
                    M_c,L_c=cls.cal_ML_simple(ST,L_s)
            while abs(L_c_need-L_c)>error_L_c:
                ST_next=ST+steps[0]
                point2=cls.cal_L_safe(ST_next,L_s)
                point1=[ST_next-steps[0],cls.cal_L_safe(ST_next-steps[0],L_s)[1]]
                k=cls.grad(point1,point2)
                if k==0:
                    raise ZeroDivisionError('k is zero.')
                else:
                    ST=(L_c_need-L_s)/k+ST
                    ST,L_c=cls.cal_L_safe(ST,L_s)
        return [[ST,L_s],[M_c,L_c]]
    
    @classmethod
    def change_const(cls,new_const):
        cls.con=new_const
        
    @classmethod
    def find_safe_point(cls,ST_init,L_s):
        L=L_s
        ST=ST_init
        ST_store=0
        judge=False
        if ST<0:
            ST=-ST
        while True:
            try:
                warnings.simplefilter('error')
                L_c=cls.cal_ML_simple(ST,L)[1]
            except:
                if judge:
                    ST=ST-ST_store
                    break
                ST/=10
                ST_store=ST
            else:
                ST+=ST_store
                judge=True
        ST_left=ST
        ST_right=ST+ST_store
        while abs(L_c)>0.003:
            ST=(ST_left+ST_right)/2
            try:
                warnings.simplefilter('error')
                L_c=cls.cal_ML_simple(ST,L)[1]
            except:
                ST_right=ST
                L_c=-10 #set a value to continue the loop
                continue
            else:
                ST_left=ST
                continue
        return ST
    
class Calculator(object):
    """Use to cal the deeper var."""
    def __init__(self):
        self.P=[]
        self.T=[]
        self.rho=[]
        self.L=[]
        self.L_init=[]
        self.m=[]
        self.M_p=[]
        self.ST=[]
        self.t=[]
        self.me=Methods()
    def cal_M_range(self,arange):
        for i in range(len(arange)):
            m=arange[i]
            self.M_p.append(m)
            self.me.con.set_M_p(m)
            self.ST.append(self.me.find_L_M([5.0,0],[0,2],[1e-8,1e-6],[1e-5,3e-3])[0][0])
            self.P.append(self.me.P)
            self.T.append(self.me.T)
            rho=[]
            for i in range(len(self.me.P)):
                rho.append(self.me.P[i]/self.me.T[i])
            self.rho.append(rho)
            self.L.append(self.me.L)
    
    def creat_t(self,arange,choose=True):
        """arange is the range of M_p.
        If choose==True, the function will run the cal_M_range. If you have run the cal_M_range, please set choose=False."""
        if choose==True:
            self.cal_M_range(arange)
        self.t.append(0.)
        n_0=self.me.con.R/self.me.con.mu
        C_p=5./2.
        for i in range(len(arange)-1):
            T_1,T_2=self.T[i][-1],self.T[i+1][-1]
            P_1,P_2=self.P[i][-1],self.P[i+1][-1]
            ST_1,ST_2=self.ST[i],self.ST[i+1]
            delta_1=C_p*(1/T_2+1/T_1)*(T_2-T_1)
            delta_2=(1/P_2+1/P_1)*(P_2-P_1)
            dt=n_0*(delta_1+delta_2)/((ST_1+ST_2)*self.me.con.Myr)
            self.t.append(self.t[i]+dt)
        
    def write_excel(self,arange,filename='Data.xlsx'):
        wb=Workbook()
        wb.create_sheet('result',index=0)
        sheet=wb[wb.sheetnames[0]]
        length=len(arange)
        sheet.append(['Information: ','range=linspace({0},{1},{2})'.format(arange[0],arange[length-1], length),'M_c={0}M_e'.format(self.me.con.M_e),'a={0}AU'.format(self.me.con.a)])
        sheet.append(['M_p/M_e','ST','L/e+24erg/s','t/Myr'])
        for i in range(len(arange)):
            sheet.append([arange[i],self.ST[i],self.L[i][0],self.t[i]])
        wb.save(filename)
    def read_excel(self,filename,sheetname='result'):
        wb=load_workbook(filename)
        sheet=wb[sheetname]
        row_max=sheet.max_row
        message=[]
        self.M_p=[]
        self.ST=[]
        self.L_init=[]
        self.t=[]
        for i in range(2,row_max,1):
            message.append(list(sheet.rows)[i])
        for i in range(len(message)):
            data=list(message[i])
            self.M_p.append(data[0].value)
            self.ST.append(data[1].value)
            self.L_init.append(data[2].value)
            self.t.append(data[3].value)
        
        
            

