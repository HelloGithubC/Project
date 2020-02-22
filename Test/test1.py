# -*- coding: utf-8 -*-
"""
Created on Wed Jun  5 20:21:57 2019

@author: 文明研究员
"""
import numpy as np
import matplotlib.pyplot as plt
from const import Const
from scipy.integrate import odeint
from openpyxl import Workbook,load_workbook
from valueDataBase import DataBase
from calculator import Calculator
import pdb

class Test(object):
    def __init__(self,a=0.3,M_c=5):
        self.con=Const(a,M_c)
        self.values=[]
        self.dValues=[]
        self.dr=0.0
        self.num=1000
        self.f=[]

    def ode_fun(self,Z,r,c_P,c_T,c_M,others,ST):
        P,T,M,L=Z
        g,alpha,beta=others
        dP=c_P*M*P/(T*r**2)
        dT=min(c_T*1e24*abs(L)*P**(alpha+1)*T**(beta-4)/M,g)*(T*dP/P)

        dM=c_M*r**2*P/T

        if dT==g*(T*dP/P):
            dL=1e-24*self.con.M_e*self.con.T_0*dM*ST

        else:
            dL=0
        return [dP,dT,dM,dL]

    
    def ode_fun_B(self,Z,r,c_P,c_T,c_M,others,ST):
        P,T,M,G,DG,L=Z
        #if L<-1000:
            #raise LOutOfRangeError('L<-1000')
        g,alpha,beta,Lambda,R_B=others

        dP=c_P*M*P/(T*r**2)
        P_0=self.con.P_0
        dT=min(c_T*1e24*abs(L)*P**(alpha+1)*T**(beta-4)/M,g)*(T*dP/P)
        dM=c_M*r**2*P/T
        dG=DG
        if P<=0.0:
            print("Error")
        try:
            dL=Lambda*7.15e-5*(dG**2+G**2/r**2)/(9e9*10**self.f(np.log10(P*P_0)-6)*R_B)
        except ValueError:
            print("Error")
        if dT==g*(T*dP/P):
            dL+=1e-24*self.con.M_e*self.con.T_0*dM*ST
        else:
            dL+=0
        ddG=6*G/r**2+(0.5*1e3*(self.f(np.log10(P*P_0)-6+1e-3)-self.f(np.log10(P*P_0)-6-1e-3))*dP/P)*dG+(r>self.con.depth)*2*self.con.M_v*(1/r-self.con.depth**2/r**3)/self.con.c**2/self.con.R_B
        
        self.values.append([P,T,M,G,L])
        #if dL>100:
            #pdb.set_trace()
        self.dValues.append([dP,dT,dM,dG,ddG,dL])
        return [dP,dT,dM,dG,ddG,dL]

    
    def find_RCB_index(self):
        P,T,M,L=self.P,self.T,self.M,self.L
        judge=(self.con.c_T*1e24*abs(L)*P**(self.con.alpha+1)*T**(self.con.beta-4)/M)<(np.ones(len(L))*self.con.g_ad)
        for i in range(len(judge)):
            if judge[i]!=judge[0]:
                self.RCB_index=i
                return i 
        return 0

    
    def cal_ML_simple(self,ST,L_s):
        c_P,c_T,c_M=self.con.cal_const()
        g,alpha,beta=self.con.g_ad,self.con.alpha,self.con.beta
        R_in,R_out=self.con.R_p,self.con.R_out
        num=self.num

        r=np.linspace(1,R_in/R_out,num)
        initial=(1.,1.,self.con.M_p,L_s)
        others=(g,alpha,beta)
        result=odeint(self.ode_fun,initial,r,args=(c_P,c_T,c_M,others,ST))
        P,T,M,L=result[:,0],result[:,1],result[:,2],result[:,3]
        self.P,self.T,self.M,self.L=P,T,M,L
        self.r=r

        self.find_RCB_index()
        return M[-1],L[-1]

    
    def cal_ML_simple_B(self,ST,L_s,B=False,G=1e-12,dG=1e-15):
        self.values=[]
        self.dValues=[]
        c_P,c_T,c_M=self.con.cal_const()
        g,alpha,beta=self.con.g_ad,self.con.alpha,self.con.beta
        R_in,R_out=self.con.R_p,self.con.R_out
        num=self.num
        base=DataBase()
        base.creat_sigma()
        self.f=base.f_sigma

        r=np.linspace(1,R_in/R_out,num)
        self.dr=r[0]-r[1]
        if B:
            initial=(1.,1.,self.con.M_p,G,dG,L_s)
        else:
            initial=(1.,1.,self.con.M_p,0.0,0.0,L_s)
        others=(g,alpha,beta,self.con.Lambda,self.con.R_B)
        result=odeint(self.ode_fun_B,initial,r,args=(c_P,c_T,c_M,others,ST))
        P,T,M,G,dg,L=result[:,0],result[:,1],result[:,2],result[:,3],result[:,4],result[:,5]
        self.P,self.T,self.M,self.L=P,T,M,L
        self.r,self.g=r,G
        self.dg=dg

        self.find_RCB_index()
        return M[-1],L[-1],G[-1]

    
    def sigma(self,P,T):
        return self.con.sigma_1*T**(3./4)*P**(-1./2)*np.exp(-self.con.sigma_2/T)

    
    def find_L_simple(self,targets, inits, steps, errors):
        """The order of targets is [M_c, L_c], of inits is [ST, L_s](Note that the unit of L_s is 1e+24 erg/s), of steps and error is the same as inits and targets."""
        """The way to find the result is one-dimention Newton iteration. First adjustion is L_s and then is ST."""
        ST,L_s=inits
        L_c_need=targets
        error_L_c=errors
        L_c=self.cal_ML_simple_B(ST,L_s,True,0.0,-1000)[1]
        i=1
        while abs(L_c_need-L_c)>error_L_c:
            ST_next=ST+steps/(2*i)
            i+=1
            L_c_next=self.cal_ML_simple_B(ST_next,L_s,True,0.0,-1000)[1]
            point1=[ST,L_c]
            point2=[ST_next,L_c_next]
            k=self.grad(point1,point2)
            if k==0:
                raise ZeroDivisionError('k is zero.Error is L_c. M_p is {0}'.format(self.con.M_p))
            else:
                ST=(L_c_need-L_c)/k+ST
                L_c=self.cal_ML_simple_B(ST,L_s,True,0.0,-1000)[1]
        return ST,L_c

    
    def find_L_two(self,targets,L_s, left,right,errors,p=False,maxTimes=20,gt=-1e+3):
        ST_left=left
        ST_right=right 
        ST=(ST_right+ST_left)/2
        L_c_need=targets
        error_L_c=errors
        L_c=self.cal_ML_simple_B(ST,L_s,True,0.0,gt)[1]
        i=0
        while abs(L_c_need-L_c)>error_L_c:
            i+=1
            if L_c>0.0:
                ST_left=ST
            else:
                ST_right=ST
            ST=(ST_right+ST_left)/2
            L_c=self.cal_ML_simple_B(ST,L_s,True,0.0,gt)[1]
            if p:
                print(ST,L_c)
            if i>maxTimes:
                break
        return ST,L_c
            
    
    def grad(self,point1,point2):
        """Return the grad of two points, so for one-dimention Newton iteration."""
        return (point2[1]-point1[1])/(point2[0]-point1[0])
    
    
    def creat_test_data(self,fileName='TestData.xlsx'):
        wb=Workbook()
        wb.create_sheet('result',index=0)
        sheet=wb[wb.sheetnames[0]]
        length=len(self.values)
        sheet.append(['Information: ','r_range=linspace({0},{1},{2})'.format(self.values[0][0],self.values[length-1][0], length),'M_c={0}M_e'.format(self.con.M_e),'M_p={0}M_e'.format(self.con.M_p),'a={0}AU'.format(self.con.a)])
        sheet.append(['r','P','T','M','G','DG','L'])
        for i in range(len(self.values)):
            sheet.append(self.values[i])
        sheet.append(['dr','dP','dT','dM','dG','dDG','dL'])
        for i in range(len(self.values)):
            sheet.append(self.dValues[i])
        wb.save(fileName)

    
    def creat_t(self,initValues,fileName='TestT.xlsx',test_start=0):
        """The order of initValues is M_p, ST, L_s"""
        M_p,ST,L_s=initValues
        P=np.ones(len(M_p)-test_start)
        T=np.ones(len(M_p)-test_start)
        L=np.ones(len(M_p)-test_start)
        for order in range(test_start,len(M_p)):
            self.con.set_M_p(M_p[order])
            ST[order]=self.find_L_simple(0.0,[ST[order],L_s[order]],1e-7,1e-5)[0]
            print("Finished: {0}/200".format(order))
            P[order]=self.P[-1]
            T[order]=self.T[-1]
            print("P,T are {0},{1}".format(P[order],T[order]))
            L[order]=self.L[0]
        self.ST=ST
        self.t=np.ones(len(M_p))
        n_0=self.con.R/self.con.mu
        C_p=5./2.
        self.t[0]=0.0
        for i in range(len(M_p)-1):
            T_1,T_2=self.T[i],self.T[i+1]
            P_1,P_2=self.P[i],self.P[i+1]
            ST_1,ST_2=self.ST[i],self.ST[i+1]
            delta_1=C_p*(1/T_2+1/T_1)*(T_2-T_1)
            delta_2=(1/P_2+1/P_1)*(P_2-P_1)
            dt=n_0*(delta_1+delta_2)/((ST_1+ST_2)*self.con.Myr)
            self.t[i+1]=self.t[i]+dt

        wb=Workbook()
        wb.create_sheet('result',index=0)
        sheet=wb[wb.sheetnames[0]]
        length=len(M_p)
        sheet.append(['Information: ','range=linspace({0},{1},{2})'.format(M_p[0],M_p[length-1], length),'M_c={0}M_e'.format(self.con.M_e),'a={0}AU'.format(self.con.a)])
        sheet.append(['M_p/M_e','ST','L/e+24erg/s','t/Myr'])
        for i in range(len(M_p)):
            sheet.append([M_p[i],self.ST[i],self.L[i],self.t[i]])
        wb.save(fileName)

    
    def find_L(self,test_start=0):
        cal=Calculator()
        cal.read_excel('Data2.xlsx')
        length=200
        P=np.ones(length-test_start)
        T=np.ones(length-test_start)
        L=np.ones(length-test_start)
        a=open('store.txt','a',encoding='utf-8')
        a.write("order,P,T,L,ST\n")
        a.close()
        for order in range(test_start,length):
            a=open('store.txt','a',encoding='utf-8')
            ST=cal.ST[order]
            L_s=cal.L_init[order]
            self.con.set_M_p(cal.M_p[order])
            ST=self.find_L_simple(0.0,[ST,L_s],1e-6,1e-5)[0]
            print("Finished: {0}/200".format(order))
            P[order]=self.P[-1]
            T[order]=self.T[-1]
            print("P,T are {0},{1}".format(P[order],T[order]))
            L[order]=self.L[0]
            a.write("{0}:{1},{2},{3},{4}\n".format(order,P[order],T[order],L[order],ST))
            a.close()


        