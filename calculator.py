from const import Const
import numpy as np
import matplotlib.pyplot as plt
import warnings
from scipy.integrate import odeint
from openpyxl import Workbook,load_workbook

class LOutOfRangeError(Exception):
    def __init__(self,message):
        Exception.__init__(self,'L<-1000')

#Can not stop.
class Methods(object):
    con=Const()    #If the Const you need is not initial, you need to change it.
    P=[]
    T=[]
    M=[]
    L=[]
    g=[]
    r=[]
    L_store=0
    test_sigma=[]
    RCB_index=0
    f=None 
    f_judge=False

    @classmethod
    def ode_fun(cls,Z,r,c_P,c_T,c_M,others,ST):
        P,T,M,L=Z
        g,alpha,beta=others
        dP=c_P*M*P/(T*r**2)
        dT=min(c_T*1e24*abs(L)*P**(alpha+1)*T**(beta-4)/M,g)*(T*dP/P)

        dM=c_M*r**2*P/T

        if dT==g*(T*dP/P):
            dL=1e-24*cls.con.M_e*cls.con.T_0*dM*ST

        else:
            dL=0
        return [dP,dT,dM,dL]

    @classmethod
    def ode_fun_B(cls,Z,r,c_P,c_T,c_M,others,ST):
        P,T,M,G,DG,L=Z
        #if L<-1000:
            #raise LOutOfRangeError('L<-1000')
        g,alpha,beta,Lambda,R_B=others

        dP=c_P*M*P/(T*r**2)
        dT=min(c_T*1e24*abs(L)*P**(alpha+1)*T**(beta-4)/M,g)*(T*dP/P)
        dM=c_M*r**2*P/T
        dG=DG
        ddG=6*G/r**2+((-0.5*dP/P)+(0.75/T+cls.con.sigma_2/T**2)*dT)*dG+(r<0.05)*2*cls.con.M_v*(1/r-0.01/r**3)*(G>0.0)
        #ddG=6*G/r**2-1/r*dG
        dL=Lambda*7.15e-5*(dG**2+G**2/r**2)/(cls.sigma(P,T)*R_B)

        if dT==g*(T*dP/P):
            dL+=1e-24*cls.con.M_e*cls.con.T_0*dM*ST
        else:
            dL+=0
        cls.test_sigma.append((-0.5*dP/P)+(0.75/T+cls.con.sigma_2/T**2)*dT)
        return [dP,dT,dM,dG,ddG,dL]

    @classmethod
    def find_RCB_index(cls):
        P,T,M,L=cls.P,cls.T,cls.M,cls.L 
        judge=(cls.con.c_T*1e24*abs(L)*P**(cls.con.alpha+1)*T**(cls.con.beta-4)/M)<(np.ones(len(L))*cls.con.g_ad)
        for i in range(len(judge)):
            if judge[i]!=judge[0]:
                cls.RCB_index=i
                return i 
        return 0

    @classmethod
    def cal_ML_simple(cls,ST,L_s):
        c_P,c_T,c_M=cls.con.cal_const()
        g,alpha,beta=cls.con.g_ad,cls.con.alpha,cls.con.beta
        R_in,R_out=cls.con.R_p,cls.con.R_out
        num=800

        r=np.linspace(1,R_in/R_out,num)
        initial=(1.,1.,cls.con.M_p,L_s)
        others=(g,alpha,beta)
        result=odeint(cls.ode_fun,initial,r,args=(c_P,c_T,c_M,others,ST))
        P,T,M,L=result[:,0],result[:,1],result[:,2],result[:,3]
        cls.P,cls.T,cls.M,cls.L=P,T,M,L
        cls.r=r

        cls.find_RCB_index()
        return M[-1],L[-1]

    @classmethod
    def cal_ML_simple_B(cls,ST,L_s,B=False,G=1e-12,dG=1e-15):
        c_P,c_T,c_M=cls.con.cal_const()
        g,alpha,beta=cls.con.g_ad,cls.con.alpha,cls.con.beta
        R_in,R_out=cls.con.R_p,cls.con.R_out
        num=800

        r=np.linspace(1,R_in/R_out,num)
        cls.test_sigma=[]
        if B:
            initial=(1.,1.,cls.con.M_p,G,dG,L_s)
        else:
            initial=(1.,1.,cls.con.M_p,0.0,0.0,L_s)
        others=(g,alpha,beta,cls.con.Lambda,cls.con.R_B)
        cls.test2=[]
        result=odeint(cls.ode_fun_B,initial,r,args=(c_P,c_T,c_M,others,ST))
        P,T,M,G,dg,L=result[:,0],result[:,1],result[:,2],result[:,3],result[:,4],result[:,5]
        cls.P,cls.T,cls.M,cls.L=P,T,M,L
        cls.r,cls.g=r,G
        cls.dg=dg

        cls.find_RCB_index()
        return M[-1],L[-1],G[-1]

    @classmethod
    def sigma(cls,P,T):
        return cls.con.sigma_1*T**(3./4)*P**(-1./2)*np.exp(-cls.con.sigma_2/T)

    @classmethod
    def insertValue()  
    @classmethod
    def cal_L_safe(cls,ST,L_s):
        try:
            L=cls.cal_ML_simple(ST,L_s)[1]
        except LOutOfRangeError:
            ST=cls.find_safe_point(ST,L_s)
            L=cls.cal_ML_simple(ST,L_s)[1]
        finally:
            return ST,L
    
    @classmethod
    def grad(cls,point1,point2):
        """Return the grad of two points, so for one-dimention Newton iteration."""
        return (point2[1]-point1[1])/(point2[0]-point1[0])
    
    
    @classmethod
    def find_L_M(cls, targets, inits, steps, errors):
        """The order of targets is [M_c, L_c], of inits is [ST, L_s](Note that the unit of L_s is 1e+24 erg/s), of steps and error is the same as inits and targets."""
        """The way to find the result is one-dimention Newton iteration. First adjustion is L_s and then is ST."""
        ST,L_s=inits
        M_c_need,L_c_need=targets
        error_M_c, error_L_c=errors
        M_c,L_c=cls.cal_ML_simple(ST,L_s)
        while abs(M_c-M_c_need)>error_M_c or abs(L_c-L_c_need)>error_L_c:
            while abs(M_c-M_c_need)>error_M_c:
                cls.error_judge=False
                point1=(L_s,M_c)
                point2=(L_s+L_s*steps[1],cls.cal_ML_simple(ST,L_s+L_s*steps[1])[0])
                k=cls.grad(point1,point2)
                if k==0:
                    raise ZeroDivisionError('k is zero.Error is M_c,M_p is {0}'.format(cls.con.M_p))
                else:
                    L_s=(M_c_need-M_c)/k+L_s
                    if L_s<0:
                        L_s=-L_s/10
                    M_c,L_c=cls.cal_ML_simple(ST,L_s)
            while abs(L_c_need-L_c)>error_L_c:
                ST_next=ST+steps[0]
                L_c_next=cls.cal_ML_simple(ST_next,L_s)[1]
                point1=[ST,L_c]
                point2=[ST_next,L_c_next]
                k=cls.grad(point1,point2)
                if k==0:
                    raise ZeroDivisionError('k is zero.Error is L_c. M_p is {0}'.format(cls.con.M_p))
                else:
                    ST=(L_c_need-L_c)/k+ST
                    ST,L_c=cls.cal_L_safe(ST,L_s)
        return [[ST,L_s],[M_c,L_c]]
    
    @classmethod
    def change_const(cls,new_const):
        cls.con=new_const
        
    @classmethod
    def find_safe_point(cls,ST_init,L_s):
        L=L_s
        ST=ST_init
        if ST<0:
            ST=-ST
        while True:
            try:
                cls.cal_ML_simple(ST,L)
            except LOutOfRangeError:
                ST/=10
            else:
                break
        return ST

class Calculator(object):
    """Use to cal the deeper var."""
    def __init__(self):
        self.P=[]
        self.T=[]
        self.rho=[]
        self.L=[]
        self.L_init=[]
        self.m=[]
        self.M_p=[]
        self.ST=[]
        self.t=[]
        self.me=Methods()
        self.g_in=0.0
    def cal_M_range(self,arange):
        initial=[0.0,1.0]
        for i in range(len(arange)):
            m=arange[i]
            if m<5.3:
                initial=[0.0,50.0]
            else:
                initial=[0.0,1.0]
            self.M_p.append(m)
            self.me.con.set_M_p(m)
            self.ST.append(self.me.find_L_M([5.0,0.0],initial,[1e-8,1e-3],[1e-5,3e-3])[0][0])
            self.P.append(self.me.P)
            self.T.append(self.me.T)
            rho=[]
            for i in range(len(self.me.P)):
                rho.append(self.me.P[i]/self.me.T[i])
            self.rho.append(rho)
            self.L.append(self.me.L)
    
    def creat_t(self,arange,choose=True):
        """arange is the range of M_p.
        If choose==True, the function will run the cal_M_range. If you have run the cal_M_range, please set choose=False."""
        if choose==True:
            self.cal_M_range(arange)
        self.t.append(0.)
        n_0=self.me.con.R/self.me.con.mu
        C_p=5./2.
        for i in range(len(arange)-1):
            T_1,T_2=self.T[i][-1],self.T[i+1][-1]
            P_1,P_2=self.P[i][-1],self.P[i+1][-1]
            ST_1,ST_2=self.ST[i],self.ST[i+1]
            delta_1=C_p*(1/T_2+1/T_1)*(T_2-T_1)
            delta_2=(1/P_2+1/P_1)*(P_2-P_1)
            dt=n_0*(delta_1+delta_2)/((ST_1+ST_2)*self.me.con.Myr)
            self.t.append(self.t[i]+dt)

    def creat_data(self,arange,data_filename='Data.xlsx',plot_filename='fig.png'):
        #arange=np.linspace(5.07,12,200) #5.07 is the min value could calculate
        self.cal_M_range(arange)
        self.creat_t(arange,False)
        plt.figure(dpi=100)
        plt.plot(self.t,arange,label=r'$M_{p}-t$')
        plt.xlabel('t/Myr')
        plt.ylabel(r'$M_{p}$')
        #plt.show()
        if type(data_filename) is bool or type(plot_filename) is bool:
            if type(data_filename) is bool:
                if data_filename:
                    self.write_excel(arange,data_filename)
            if type(plot_filename) is bool:
                if plot_filename:
                    plt.savefig(plot_filename)
        else:
            plt.savefig(plot_filename)
            self.write_excel(arange,data_filename)

    def cal_g_in(self,m,ST_init,L_s,g,L_in_target,ST_step=1e-5,error=1e-5):
        self.me.con.set_M_p(m)
        ST=ST_init
        L_in,g_in=self.me.cal_ML_simple_B(ST,L_s,True,g)[1:3]
        while abs(L_in-L_in_target)>error:
            ST_next=ST+ST_step
            L_in_next=self.me.cal_ML_simple_B(ST_next,L_s,True,g)[1]
            k=self.me.grad([ST,L_in],[ST_next,L_in_next])
            if k==0:
                raise ZeroDivisionError('k is zero.')
            ST=(L_in_target-L_in)/k+ST 
            L_in,g_in=self.me.cal_ML_simple_B(ST,L_s,True,g)[1:3]
        return ST,g_in

        
    def write_excel(self,arange,filename='Data.xlsx'):
        wb=Workbook()
        wb.create_sheet('result',index=0)
        sheet=wb[wb.sheetnames[0]]
        length=len(arange)
        sheet.append(['Information: ','range=linspace({0},{1},{2})'.format(arange[0],arange[length-1], length),'M_c={0}M_e'.format(self.me.con.M_e),'a={0}AU'.format(self.me.con.a)])
        sheet.append(['M_p/M_e','ST','L/e+24erg/s','t/Myr'])
        for i in range(len(arange)):
            sheet.append([arange[i],self.ST[i],self.L[i][0],self.t[i]])
        wb.save(filename)
    def read_excel(self,filename,sheetname='result'):
        wb=load_workbook(filename)
        sheet=wb[sheetname]
        row_max=sheet.max_row
        message=[]
        self.M_p=[]
        self.ST=[]
        self.L_init=[]
        self.t=[]
        for i in range(2,row_max,1):
            message.append(list(sheet.rows)[i])
        for i in range(len(message)):
            data=list(message[i])
            self.M_p.append(data[0].value)
            self.ST.append(data[1].value)
            self.L_init.append(data[2].value)
            self.t.append(data[3].value)
        
        
            

